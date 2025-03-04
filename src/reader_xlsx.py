import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
class ExcelDataReader: #Class to declare the variables who get the file path and file name
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def excel_data(self): # Here get the excel data
        self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
        workbook = load_workbook(self.file_path)
        self.sheet = workbook[self.sheet_name]

    def get_excel_row_data(self):
        # Inicializar a lista
        row_data_list = []
        # Começar na linha 2
        for row in range(2, self.sheet.max_row + 1):  # Iterar sobre todas as linhas
            # Obter o valor da coluna F
            cell_value = self.sheet[f'F{row}'].value
            if cell_value is not None and (isinstance(cell_value, str) and cell_value.strip() or isinstance(cell_value, datetime)):  # Verificar se a célula não está vazia e é uma string ou uma data
                row_data_object = RowData(self.sheet, row)
                row_data_list.append(row_data_object)
        return row_data_list

class RowData: # Here declare the variables who will be defined about the row
    def __init__(self, sheet, row_idx): 
        self.data_demissao                      = sheet[f'F{row_idx}'].value
        self.setor                              = sheet[f'G{row_idx}'].value
        self.cargo                              = sheet[f'H{row_idx}'].value
        self.iniciativa_desligamento            = sheet[f'I{row_idx}'].value
        self.motivodesliga                      = sheet[f'J{row_idx}'].value 
        self.avaliacaojornada                   = sheet[f'K{row_idx}'].value
        self.dms_consideracoes_01               = sheet[f'L{row_idx}'].value 
        self.beneficios                         = sheet[f'M{row_idx}'].value
        self.beneficiostxt                      = sheet[f'N{row_idx}'].value
        self.beneficiosrelev                    = sheet[f'O{row_idx}'].value
        self.avalia_ambiente_fgm                = sheet[f'P{row_idx}'].value
        self.sugestaoMudancas                   = sheet[f'Q{row_idx}'].value
        self.feedback                           = sheet[f'R{row_idx}'].value
        self.comunicacao                        = sheet[f'S{row_idx}'].value
        self.comunicatxt                        = sheet[f'T{row_idx}'].value
        self.treinamentos                       = sheet[f'U{row_idx}'].value
        self.oportunidades                      = sheet[f'V{row_idx}'].value
        self.buscarlideranca                    = sheet[f'W{row_idx}'].value
        self.gestaoacessibil                    = sheet[f'X{row_idx}'].value
        self.treinado                           = sheet[f'Y{row_idx}'].value
        self.indicaria_fgm                      = sheet[f'Z{row_idx}'].value
        self.recomendatxt                       = sheet[f'AA{row_idx}'].value
        self.mensagemparafgm                    = sheet[f'AB{row_idx}'].value